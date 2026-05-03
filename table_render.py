from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


DEFAULT_FONT_SIZE = 16
MIN_COL_WIDTH = 90
MAX_COL_WIDTH = 260
ROW_HEIGHT = 34
PADDING = 8


def render_xlsx_region(
    file_path: str | Path,
    sheet_name: str,
    row_start: int,
    row_end: int,
    output_path: str | Path,
    col_start: int = 1,
    col_end: int | None = None,
    highlight_row: int | None = None,
) -> Path:
    path = Path(file_path)
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("表格截图目前只支持 .xlsx/.xlsm；.xls 请先另存为 .xlsx")

    workbook = load_workbook(path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet 不存在：{sheet_name}")
    sheet = workbook[sheet_name]
    row_start = max(1, row_start)
    row_end = min(sheet.max_row, row_end)
    col_end = col_end or min(sheet.max_column, 12)

    col_widths = compute_col_widths(sheet, row_start, row_end, col_start, col_end)
    image_width = sum(col_widths) + 1
    image_height = (row_end - row_start + 1) * ROW_HEIGHT + 1
    image = Image.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)
    font = load_font()

    y = 0
    for row in range(row_start, row_end + 1):
        x = 0
        for offset, col in enumerate(range(col_start, col_end + 1)):
            width = col_widths[offset]
            cell = sheet.cell(row=row, column=col)
            fill = cell_fill_rgb(cell) or ("#FFF2CC" if highlight_row == row else "#FFFFFF")
            draw.rectangle([x, y, x + width, y + ROW_HEIGHT], fill=fill, outline="#808080")
            text = format_cell(cell.value)
            if text:
                draw.text((x + PADDING, y + 7), trim_text(text, width, font), fill="#111111", font=font)
            x += width
        y += ROW_HEIGHT

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    image.save(output)
    return output


def compute_col_widths(sheet, row_start: int, row_end: int, col_start: int, col_end: int) -> list[int]:
    widths = []
    for col in range(col_start, col_end + 1):
        letter = get_column_letter(col)
        configured = sheet.column_dimensions[letter].width
        if configured:
            width = int(configured * 8)
        else:
            max_len = 8
            for row in range(row_start, row_end + 1):
                max_len = max(max_len, len(format_cell(sheet.cell(row=row, column=col).value)))
            width = max_len * 12
        widths.append(max(MIN_COL_WIDTH, min(MAX_COL_WIDTH, width)))
    return widths


def cell_fill_rgb(cell) -> str | None:
    fill = cell.fill
    if not fill or fill.fill_type != "solid":
        return None
    color = fill.fgColor
    if color.type == "rgb" and color.rgb:
        rgb = color.rgb[-6:]
        if rgb.upper() != "000000":
            return f"#{rgb}"
    return None


def load_font():
    candidates = [
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf",
        "C:/Windows/Fonts/arial.ttf",
    ]
    for candidate in candidates:
        if Path(candidate).exists():
            return ImageFont.truetype(candidate, DEFAULT_FONT_SIZE)
    return ImageFont.load_default()


def trim_text(text: str, width: int, font) -> str:
    limit = max(4, int((width - PADDING * 2) / 9))
    text = text.replace("\n", " ")
    return text if len(text) <= limit else text[: limit - 1] + "…"


def format_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()
