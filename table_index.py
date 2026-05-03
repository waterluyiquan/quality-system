from __future__ import annotations

import hashlib
import json
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


SUPPORTED_TABLE_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
MAX_ROWS_PER_SHEET = 3000
MAX_COLS_PER_SHEET = 80


@dataclass
class TableRow:
    path: str
    file: str
    sheet: str
    row: int
    text: str
    cells_json: str
    file_type: str


def default_table_db(db_dir: str | Path) -> Path:
    db_path = Path(db_dir) / "table_index.sqlite"
    db_path.parent.mkdir(parents=True, exist_ok=True)
    return db_path


def connect(db_path: str | Path) -> sqlite3.Connection:
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    return conn


def init_db(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS table_rows (
            id TEXT PRIMARY KEY,
            path TEXT NOT NULL,
            file TEXT NOT NULL,
            sheet TEXT NOT NULL,
            row INTEGER NOT NULL,
            text TEXT NOT NULL,
            cells_json TEXT NOT NULL,
            file_type TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE VIRTUAL TABLE IF NOT EXISTS table_rows_fts
        USING fts5(text, file, sheet, content='table_rows', content_rowid='rowid')
        """
    )
    conn.commit()


def rebuild_table_index(source_dir: str | Path, db_dir: str | Path, progress_callback=None) -> int:
    db_path = default_table_db(db_dir)
    conn = connect(db_path)
    init_db(conn)
    conn.execute("DELETE FROM table_rows_fts")
    conn.execute("DELETE FROM table_rows")
    conn.commit()

    files = list(iter_table_files(Path(source_dir)))
    total_rows = 0
    for index, path in enumerate(files, start=1):
        if progress_callback:
            progress_callback(index, len(files), path)
        row_count = 0
        for row in extract_table_rows(path):
            row_id = make_row_id(row)
            cursor = conn.execute(
                """
                INSERT INTO table_rows(id, path, file, sheet, row, text, cells_json, file_type)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (row_id, row.path, row.file, row.sheet, row.row, row.text, row.cells_json, row.file_type),
            )
            rowid = cursor.lastrowid
            conn.execute(
                "INSERT INTO table_rows_fts(rowid, text, file, sheet) VALUES (?, ?, ?, ?)",
                (rowid, row.text, row.file, row.sheet),
            )
            row_count += 1
        total_rows += row_count
        conn.commit()
    conn.close()
    return total_rows


def search_table_index(db_dir: str | Path, query: str, limit: int = 20, context_rows: int = 2) -> list[dict]:
    db_path = default_table_db(db_dir)
    conn = connect(db_path)
    init_db(conn)
    sql_query = build_fts_query(query)
    try:
        rows = conn.execute(
            """
            SELECT tr.*, bm25(table_rows_fts) AS score
            FROM table_rows_fts
            JOIN table_rows tr ON tr.rowid = table_rows_fts.rowid
            WHERE table_rows_fts MATCH ?
            ORDER BY score
            LIMIT ?
            """,
            (sql_query, limit),
        ).fetchall()
    except sqlite3.OperationalError:
        rows = conn.execute(
            """
            SELECT tr.*, 0 AS score
            FROM table_rows tr
            WHERE tr.text LIKE ?
            LIMIT ?
            """,
            (f"%{query}%", limit),
        ).fetchall()

    results = [row_to_result(row, context_rows=context_rows) for row in rows]
    conn.close()
    return results


def row_to_result(row: sqlite3.Row, context_rows: int = 2) -> dict:
    row_number = int(row["row"])
    return {
        "path": row["path"],
        "file": row["file"],
        "sheet": row["sheet"],
        "row": row_number,
        "row_start": max(1, row_number - context_rows),
        "row_end": row_number + context_rows,
        "text": row["text"],
        "cells": json.loads(row["cells_json"]),
        "file_type": row["file_type"],
        "score": row["score"],
    }


def iter_table_files(source_dir: Path) -> Iterable[Path]:
    for path in sorted(source_dir.rglob("*")):
        if path.is_file() and path.suffix.lower() in SUPPORTED_TABLE_EXTENSIONS and not path.name.startswith("~$"):
            yield path


def extract_table_rows(path: Path) -> Iterable[TableRow]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        yield from extract_xlsx_rows(path)
    elif suffix == ".xls":
        yield from extract_xls_rows(path)


def extract_xlsx_rows(path: Path) -> Iterable[TableRow]:
    suffix = path.suffix.lower()
    workbook = load_workbook(path, data_only=True, read_only=True)
    for sheet in workbook.worksheets:
        max_row = min(sheet.max_row or 0, MAX_ROWS_PER_SHEET)
        max_col = min(sheet.max_column or 0, MAX_COLS_PER_SHEET)
        for row_index, row in enumerate(sheet.iter_rows(max_row=max_row, max_col=max_col), start=1):
            if not row:
                continue
            values = []
            cells = []
            for col_index, cell in enumerate(row, start=1):
                text = format_value(cell.value)
                if text:
                    values.append(text)
                    cells.append({"col": getattr(cell, "column", col_index), "value": text})
            if values:
                yield TableRow(
                    path=str(path),
                    file=path.name,
                    sheet=sheet.title,
                    row=getattr(row[0], "row", row_index),
                    text=" | ".join(values),
                    cells_json=json.dumps(cells, ensure_ascii=False),
                    file_type=suffix.lstrip("."),
                )


def extract_xls_rows(path: Path) -> Iterable[TableRow]:
    import xlrd

    workbook = xlrd.open_workbook(path, on_demand=True)
    for sheet in workbook.sheets():
        for row_index in range(min(sheet.nrows, MAX_ROWS_PER_SHEET)):
            values = []
            cells = []
            for col_index in range(min(sheet.ncols, MAX_COLS_PER_SHEET)):
                text = format_value(sheet.cell_value(row_index, col_index))
                if text:
                    values.append(text)
                    cells.append({"col": col_index + 1, "value": text})
            if values:
                yield TableRow(
                    path=str(path),
                    file=path.name,
                    sheet=sheet.name,
                    row=row_index + 1,
                    text=" | ".join(values),
                    cells_json=json.dumps(cells, ensure_ascii=False),
                    file_type="xls",
                )


def format_value(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        try:
            return str(int(float(text)))
        except ValueError:
            return text
    return text


def build_fts_query(query: str) -> str:
    tokens = [token.strip() for token in query.replace('"', " ").split() if token.strip()]
    if not tokens:
        return query
    return " OR ".join(f'"{token}"' for token in tokens)


def make_row_id(row: TableRow) -> str:
    raw = f"{row.path}|{row.sheet}|{row.row}|{row.text}"
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()
