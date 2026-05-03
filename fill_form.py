from __future__ import annotations

from io import BytesIO
from pathlib import Path
from dataclasses import dataclass
from datetime import datetime
import json
import re

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

from rag import extract_value, format_source, search


EVIDENCE_SHEET = "填写依据"
FILLABLE_EXTENSIONS = {".xlsx", ".xlsm"}
MANIFEST_NAME = "batch_manifest.json"
HEADER_KEYWORDS = ("字段", "项目", "名称", "指标", "参数", "内容")
VALUE_KEYWORDS = ("填写", "结果", "答案", "值", "说明", "备注")


@dataclass
class FillTarget:
    sheet_name: str
    row: int
    field_col: int
    value_col: int
    field_name: str


def fill_excel(
    input_file,
    output_path: str | Path | None = None,
    mode: str = "auto",
    overwrite: bool = False,
    max_fields: int | None = None,
) -> tuple[bytes, list[dict[str, str]]]:
    workbook = load_workbook(input_file)
    evidence = reset_evidence_sheet(workbook)
    evidence.append(["Sheet", "单元格", "字段名", "填写值", "来源文件", "位置", "原文摘录", "状态"])

    results: list[dict[str, str]] = []
    targets = iter_fill_targets(workbook, mode=mode)
    if max_fields:
        targets = targets[:max_fields]
    for target in targets:
        sheet = workbook[target.sheet_name]
        value_cell = sheet.cell(row=target.row, column=target.value_col)
        if isinstance(value_cell, MergedCell):
            continue
        if value_cell.value not in (None, "") and not overwrite:
            continue

        hits = search(f"{target.field_name} 是什么？", top_k=4)
        value, status = extract_value(target.field_name, hits)
        if value:
            value_cell.value = value

        top_hit = hits[0] if hits else {"metadata": {}, "text": ""}
        source = top_hit.get("metadata", {}).get("file", "")
        location = format_source(top_hit.get("metadata", {})) if hits else ""
        excerpt = (top_hit.get("text") or "")[:300]
        cell_ref = value_cell.coordinate
        evidence.append([target.sheet_name, cell_ref, target.field_name, value, source, location, excerpt, status])
        results.append(
            {
                "Sheet": target.sheet_name,
                "单元格": cell_ref,
                "字段名": target.field_name,
                "填写值": value,
                "来源文件": source,
                "位置": location,
                "原文摘录": excerpt,
                "状态": status,
            }
        )

    buffer = BytesIO()
    workbook.save(buffer)
    content = buffer.getvalue()
    if output_path:
        Path(output_path).parent.mkdir(exist_ok=True)
        Path(output_path).write_bytes(content)
    return content, results


def iter_fill_targets(workbook, mode: str = "auto") -> list[FillTarget]:
    targets: list[FillTarget] = []
    sheets = [sheet for sheet in workbook.worksheets if sheet.title != EVIDENCE_SHEET]
    for sheet in sheets:
        if mode == "fixed":
            targets.extend(iter_fixed_targets(sheet))
        else:
            auto_targets = iter_auto_targets(sheet)
            targets.extend(auto_targets or iter_fixed_targets(sheet))
    return targets


def iter_fixed_targets(sheet) -> list[FillTarget]:
    targets = []
    for row in range(1, sheet.max_row + 1):
        field_name = cell_text(sheet.cell(row=row, column=1).value)
        if field_name:
            targets.append(FillTarget(sheet.title, row, 1, 2, field_name))
    return targets


def iter_auto_targets(sheet) -> list[FillTarget]:
    header_row, field_col, value_col = detect_header_columns(sheet)
    if header_row and field_col and value_col:
        return [
            FillTarget(sheet.title, row, field_col, value_col, field_name)
            for row in range(header_row + 1, sheet.max_row + 1)
            if (field_name := cell_text(sheet.cell(row=row, column=field_col).value))
        ]

    targets: list[FillTarget] = []
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column):
            field_name = cell_text(sheet.cell(row=row, column=col).value)
            next_cell = sheet.cell(row=row, column=col + 1)
            if next_cell.value not in (None, ""):
                continue
            if looks_like_field_name(field_name) and not isinstance(next_cell, MergedCell):
                targets.append(FillTarget(sheet.title, row, col, col + 1, field_name))
    return dedupe_targets(targets)


def detect_header_columns(sheet) -> tuple[int | None, int | None, int | None]:
    max_header_row = min(sheet.max_row, 12)
    for row in range(1, max_header_row + 1):
        headers = {col: cell_text(sheet.cell(row=row, column=col).value) for col in range(1, sheet.max_column + 1)}
        field_col = first_matching_col(headers, HEADER_KEYWORDS)
        value_col = first_matching_col(headers, VALUE_KEYWORDS, exclude={field_col})
        if field_col and value_col:
            return row, field_col, value_col
    return None, None, None


def first_matching_col(headers: dict[int, str], keywords: tuple[str, ...], exclude: set[int | None] | None = None) -> int | None:
    exclude = exclude or set()
    for col, header in headers.items():
        if col in exclude:
            continue
        if any(keyword in header for keyword in keywords):
            return col
    return None


def looks_like_field_name(text: str) -> bool:
    if not text or len(text) > 80:
        return False
    if text.endswith(("：", ":")):
        return True
    return any(keyword in text for keyword in HEADER_KEYWORDS)


def cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip().strip(":：")


def dedupe_targets(targets: list[FillTarget]) -> list[FillTarget]:
    seen = set()
    unique = []
    for target in targets:
        key = (target.sheet_name, target.row, target.value_col)
        if key in seen:
            continue
        seen.add(key)
        unique.append(target)
    return unique


def reset_evidence_sheet(workbook):
    if EVIDENCE_SHEET in workbook.sheetnames:
        del workbook[EVIDENCE_SHEET]
    return workbook.create_sheet(EVIDENCE_SHEET)


def batch_fill_folder(
    source_dir: str | Path,
    export_dir: str | Path,
    mode: str = "auto",
    overwrite: bool = False,
    max_files: int | None = None,
    max_fields_per_file: int | None = 30,
    skip_done: bool = True,
    progress_callback=None,
) -> list[dict[str, str | int]]:
    source_path = Path(source_dir)
    export_path = Path(export_dir)
    export_path.mkdir(parents=True, exist_ok=True)
    results: list[dict[str, str | int]] = []
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    manifest = load_batch_manifest(export_path)

    files = list(iter_excel_files(source_path))
    if skip_done:
        files = [path for path in files if file_signature(path) not in manifest.get("processed", {})]
    if max_files:
        files = files[:max_files]

    for index, path in enumerate(files, start=1):
        if progress_callback:
            progress_callback(index, len(files), path)
        signature = file_signature(path)
        source_for_fill = path
        converted_path = ""
        if path.suffix.lower() not in FILLABLE_EXTENSIONS:
            try:
                source_for_fill = convert_legacy_excel_to_xlsx(path, source_path, export_path)
                converted_path = str(source_for_fill)
            except Exception as exc:
                row = {"文件": str(path), "状态": "失败", "原因": f".xls 自动另存为 .xlsx 失败：{exc}", "填写字段数": 0}
                results.append(row)
                record_batch_result(manifest, signature, row)
                save_batch_manifest(export_path, manifest)
                continue

        relative = path.relative_to(source_path)
        output_name = f"filled_{timestamp}_{relative.with_suffix('.xlsx').name}"
        output_path = export_path / relative.parent / output_name
        try:
            _, rows = fill_excel(
                source_for_fill,
                output_path=output_path,
                mode=mode,
                overwrite=overwrite,
                max_fields=max_fields_per_file,
            )
            row = {"文件": str(path), "状态": "完成", "原因": "", "填写字段数": len(rows), "输出文件": str(output_path), "转换后文件": converted_path}
            results.append(row)
            record_batch_result(manifest, signature, row)
        except Exception as exc:
            row = {"文件": str(path), "状态": "失败", "原因": str(exc), "填写字段数": 0, "转换后文件": converted_path}
            results.append(row)
            record_batch_result(manifest, signature, row)
        save_batch_manifest(export_path, manifest)

    return results


def iter_excel_files(source_path: Path) -> list[Path]:
    files = []
    for path in sorted(source_path.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith("~$") or path.name.startswith("filled_"):
            continue
        if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            files.append(path)
    return files


def pending_excel_files(source_path: Path, export_path: Path) -> list[Path]:
    manifest = load_batch_manifest(export_path)
    processed = manifest.get("processed", {})
    return [path for path in iter_excel_files(source_path) if file_signature(path) not in processed]


def load_batch_manifest(export_path: Path) -> dict:
    path = export_path / MANIFEST_NAME
    if not path.exists():
        return {"processed": {}}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {"processed": {}}


def save_batch_manifest(export_path: Path, manifest: dict) -> None:
    path = export_path / MANIFEST_NAME
    path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def record_batch_result(manifest: dict, signature: str, row: dict[str, str | int]) -> None:
    manifest.setdefault("processed", {})[signature] = {
        "time": datetime.now().isoformat(timespec="seconds"),
        "result": row,
    }


def reset_batch_manifest(export_path: str | Path) -> None:
    path = Path(export_path) / MANIFEST_NAME
    if path.exists():
        path.unlink()


def file_signature(path: Path) -> str:
    stat = path.stat()
    return f"{path.resolve()}|{stat.st_size}|{int(stat.st_mtime)}"


def convert_legacy_excel_to_xlsx(path: Path, source_path: Path, export_path: Path) -> Path:
    relative = path.relative_to(source_path)
    target_dir = export_path / "_converted_xls" / relative.parent
    target_dir.mkdir(parents=True, exist_ok=True)
    target = target_dir / f"{safe_stem(path.stem)}.xlsx"
    if target.exists() and target.stat().st_mtime >= path.stat().st_mtime:
        return target

    try:
        import win32com.client
    except Exception as exc:
        raise RuntimeError("需要安装 pywin32，并且本机需要有 Microsoft Excel") from exc

    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(path.resolve()))
        workbook.SaveAs(str(target.resolve()), FileFormat=51)
        return target
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()


def safe_stem(stem: str) -> str:
    return re.sub(r'[<>:"/\\|?*]+', "_", stem).strip() or "converted"
